import csv
import io
import re
from typing import List, Dict, Any

import requests
from django.core.files.uploadedfile import UploadedFile
from rest_framework.exceptions import ValidationError

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

HEADER_PATTERN = re.compile(r"[\s\-]+")


def _normalize_header(value: str) -> str:
    return HEADER_PATTERN.sub("_", value.strip().lower())


def _rows_from_csv(data: str) -> List[Dict[str, Any]]:
    if not data:
        return []
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(data[:1024])
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(data), dialect=dialect)
    reader.fieldnames = [
        _normalize_header(h or "") for h in (reader.fieldnames or [])
    ]
    rows: List[Dict[str, Any]] = []
    for row in reader:
        if not row:
            continue
        if all((value in (None, "", " ") for value in row.values())):
            continue
        rows.append({k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()})
    return rows


def _rows_from_xlsx(file: UploadedFile) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise ValidationError("openpyxl is required to handle .xlsx files")
    workbook = load_workbook(filename=file, read_only=True, data_only=True)
    sheet = workbook.active
    headers = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        value = cell.value or ""
        headers.append(_normalize_header(str(value)))
    rows: List[Dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "", " ") for value in row):
            continue
        record = {
            headers[idx]: (str(value).strip() if isinstance(value, str) else value)
            for idx, value in enumerate(row)
            if idx < len(headers)
        }
        rows.append(record)
    workbook.close()
    return rows


def _load_file(file: UploadedFile) -> List[Dict[str, Any]]:
    name = (file.name or "").lower()
    if name.endswith(".csv"):
        raw = file.read()
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        return _rows_from_csv(text)
    if name.endswith(".xlsx"):
        return _rows_from_xlsx(file)
    raise ValidationError("Unsupported file type. Please upload a CSV or XLSX file.")


def _load_sheet_url(url: str) -> List[Dict[str, Any]]:
    url = url.strip()
    if not url:
        raise ValidationError("sheet_url cannot be empty")
    if "/edit" in url:
        url = url.split("/edit", 1)[0] + "/export?format=csv"
    elif "?" not in url:
        if url.endswith("/"):
            url = url[:-1]
        url = f"{url}/gviz/tq?tqx=out:csv"
    response = requests.get(url, timeout=15)
    response.raise_for_status()
    text = response.content.decode("utf-8")
    return _rows_from_csv(text)


def load_tabular(request) -> List[Dict[str, Any]]:
    if request.FILES.get("file"):
        return _load_file(request.FILES["file"])

    sheet_url = request.data.get("sheet_url") or request.query_params.get("sheet_url")
    if sheet_url:
        return _load_sheet_url(sheet_url)

    raise ValidationError("Provide a CSV/XLSX file or a sheet_url.")
