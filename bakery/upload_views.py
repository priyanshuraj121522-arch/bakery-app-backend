import csv
import io
from typing import Dict, List, Optional

# --- UPLOAD UPGRADE START ---
from django.shortcuts import get_object_or_404
from django_q.tasks import async_task, result
from django_q.models import Task
# --- UPLOAD UPGRADE END ---

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import UploadTask

# --- UPLOAD UPGRADE START ---
MAX_UPLOAD_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_ROWS = 50_000

try:
    import pandas as pd  # type: ignore
except ImportError:  # pragma: no cover - handled below
    pd = None  # type: ignore

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover - handled below
    load_workbook = None  # type: ignore


REQUIRED_COLUMNS: Dict[str, List[str]] = {
    "sales": ["billed_at", "total"],
    "products": ["sku", "name", "price"],
    "inventory": ["product_sku", "qty", "unit_cost"],
    "kitchen": ["ingredient", "qty"],
}


def _unsupported_response():
    return Response(
        {"error": "Unsupported file type. Upload .csv or .xlsx files."},
        status=status.HTTP_400_BAD_REQUEST,
    )


def _parse_headers_and_rowcount(file_bytes: bytes, lowered_name: str):
    if lowered_name.endswith(".csv"):
        text_stream = io.StringIO(file_bytes.decode("utf-8-sig", errors="ignore"))
        reader = csv.reader(text_stream)
        try:
            headers = next(reader)
        except StopIteration:
            return [], 0
        row_count = sum(1 for _ in reader)
        return headers, row_count
    if lowered_name.endswith(".xlsx"):
        if load_workbook is None:
            raise ValueError("Server missing openpyxl")
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            iterator = workbook.active.iter_rows(min_row=1, max_row=1, values_only=True)
            first_row = next(iterator, None)
            headers = [str(cell).strip() if cell is not None else "" for cell in first_row] if first_row else []
            row_count = max(workbook.active.max_row - 1, 0)
        finally:
            workbook.close()
        return headers, row_count
    raise ValueError("Unsupported file type")


def _task_state(task_id: str):
    res = result(task_id)
    if res is not None:
        return {"state": "SUCCESS", "result": res}
    t = Task.objects.filter(id=task_id).only("success", "started", "stopped").first()
    if not t:
        return {"state": "PENDING"}
    if t.success is True and t.stopped:
        return {"state": "SUCCESS"}
    if t.success is False and t.stopped:
        return {"state": "FAILURE"}
    if t.started and not t.stopped:
        return {"state": "RUNNING"}
    return {"state": "QUEUED"}


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_data(request):
    if pd is None:
        return Response({"error": "Server missing pandas/openpyxl"}, status=status.HTTP_400_BAD_REQUEST)

    upload_type = request.query_params.get("type")
    if upload_type not in REQUIRED_COLUMNS:
        return Response(
            {"error": "Invalid type parameter. Expected one of sales, products, inventory, kitchen."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    upload = request.FILES.get("file")
    if not upload:
        return Response({"error": "File is required."}, status=status.HTTP_400_BAD_REQUEST)

    name = upload.name or ""
    lowered = name.lower()

    size = getattr(upload, "size", None)
    if size is None:
        current_pos = upload.tell()
        upload.seek(0, io.SEEK_END)
        size = upload.tell()
        upload.seek(current_pos)
    if size and size > MAX_UPLOAD_SIZE_BYTES:
        return Response(
            {"error": "File too large. Maximum allowed size is 10 MB."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    file_bytes = upload.read()
    if len(file_bytes) > MAX_UPLOAD_SIZE_BYTES:
        return Response(
            {"error": "File too large. Maximum allowed size is 10 MB."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        headers, row_count = _parse_headers_and_rowcount(file_bytes, lowered)
    except ValueError as exc:
        message = str(exc)
        if "Unsupported" in message:
            return _unsupported_response()
        return Response({"error": message}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as exc:
        return Response({"error": f"Failed to inspect file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

    required = REQUIRED_COLUMNS[upload_type]
    missing = [col for col in required if col not in headers]
    if missing:
        return Response(
            {"error": f"Missing required columns: {', '.join(missing)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if row_count > MAX_ROWS:
        return Response(
            {"error": f"Too many rows ({row_count}). Maximum allowed is {MAX_ROWS}."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    task = UploadTask.objects.create(filename=name or "upload")

    try:
        task_id = async_task(
            process_upload_task,
            task.pk,
            upload_type,
            file_bytes,
            name,
            q_options={"task_id": str(task.pk)},
        )
    except Exception as exc:
        task.status = UploadTask.STATUS_FAILED
        task.save(update_fields=["status"])
        return Response(
            {"error": f"Failed to queue upload task: {exc}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    return Response(
        {
            "status": "queued",
            "task_id": str(task_id),
            "type": upload_type,
            "detail": "File accepted. Processing in background.",
        },
        status=status.HTTP_202_ACCEPTED,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def upload_status(request, pk: int):
    task = get_object_or_404(UploadTask, pk=pk)

    payload: Dict[str, Optional[object]] = {
        "task_id": str(task.pk),
        "status": task.status,
        "created_at": task.created_at.isoformat(),
    }

    state_info = _task_state(str(task.pk))
    if state_info:
        payload["job"] = state_info.get("state")
        result_payload = state_info.get("result")
        if isinstance(result_payload, dict):
            for key in ("rows", "preview", "error", "type"):
                if key in result_payload:
                    payload[key] = result_payload[key]
        if state_info.get("state") == "FAILURE" and "error" not in payload:
            payload["error"] = "Processing failed."

    return Response(payload, status=status.HTTP_200_OK)


def process_upload_task(task_id: int, upload_type: str, file_bytes: bytes, filename: str):
    if pd is None:
        return

    task = UploadTask.objects.filter(pk=task_id).first()
    if not task:
        return

    task.status = UploadTask.STATUS_RUNNING
    task.save(update_fields=["status"])

    try:
        buffer = io.BytesIO(file_bytes)
        if filename.lower().endswith(".csv"):
            df = pd.read_csv(buffer)
        elif filename.lower().endswith(".xlsx"):
            df = pd.read_excel(buffer)
        else:
            raise ValueError("Unsupported file type for background processing.")

        required = REQUIRED_COLUMNS[upload_type]
        missing = [col for col in required if col not in df.columns]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

        total_rows = int(len(df))
        if total_rows > MAX_ROWS:
            raise ValueError(f"Too many rows ({total_rows}). Maximum allowed is {MAX_ROWS}.")

        preview: Optional[List[Dict]] = df.head(3).to_dict(orient="records") if total_rows else []

        task.status = UploadTask.STATUS_DONE
        task.save(update_fields=["status"])
        return {"rows": total_rows, "preview": preview}
    except Exception as exc:
        task.status = UploadTask.STATUS_FAILED
        task.save(update_fields=["status"])
        raise
# --- UPLOAD UPGRADE END ---
